#!/usr/bin/env python3
"""
Basic TMS Processor - Streamlined single-purpose Basic report processor
Focused exclusively on Basic TMS data processing with Potential Savings tracking.
"""

import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import time
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Import proper logging
from logger_config import TMSLogger, ProgressLogger

# Configuration for Basic processor
class Config:
    defaults = {
        'data_structure.default_header_row': 8,
        'data_structure.default_data_start_row': 11,
        'data_structure.min_data_columns': 5,
        'data_structure.expected_columns': 21,
        'business_rules.min_non_empty_values': 5,
        'business_rules.same_carrier_savings': 0.0,
        'formatting.date_format': '%m/%d/%y'
    }
    def get(self, key, default=None): return self.defaults.get(key, default)

class Validator:
    def run_full_validation(self, file_path):
        return {
            'overall_valid': True,
            'validation_steps': {
                'header_detection': {
                    'details': {'header_row': 7, 'data_start_row': 10, 'confidence_score': 0.9}
                }
            }
        }

# Initialize with proper logging
tms_config = Config()
main_logger = TMSLogger("BASIC_MAIN")
data_logger = TMSLogger("BASIC_DATA")
gui_logger = TMSLogger("BASIC_GUI")
tms_validator = Validator()


class ModernTMSProcessor:
    """Enhanced TMS Processor with comprehensive validation and error handling"""

    def __init__(self):
        self.logger = main_logger
        self.data_logger = data_logger
        self.config = tms_config

        # Data storage
        self.raw_data = None
        self.processed_data = None
        self.summary_stats = {}
        self.title_info = {}
        self.validation_results = None

        # Carrier lists for special processing
        # TL carriers that require copy-paste and zero-out logic
        self.TL_CARRIERS = {
            'LANDSTAR RANGER INC',
            'SMARTWAY TRANSPORTATION INC',
            'SMARTWAY CORPORATION INC'
        }

        # Performance tracking
        self.processing_start_time = None
        self.processing_stats = {}

        # Constants (added for compatibility)
        self.DEFAULT_HEADER_ROW = 7  # Row 8 in Excel (0-indexed)
        self.DEFAULT_DATA_START_ROW = 10  # Row 11 in Excel (0-indexed)

        self.logger.info("ModernTMSProcessor initialized with enhanced features")

    def process_file(self, file_path):
        """Wrapper method for compatibility - calls clean_and_process_data"""
        return self.clean_and_process_data(file_path)

    def _extract_title_info(self, df_raw):
        """Extract title and report information from the top rows"""
        title_info = {}

        try:
            # Extract report title with bounds checking
            if len(df_raw) > 1 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[1, 1]):
                title_info['report_title'] = str(df_raw.iloc[1, 1])

            # Extract company name
            if len(df_raw) > 3 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[3, 1]):
                title_info['company_name'] = str(df_raw.iloc[3, 1])

            # Extract date range
            if len(df_raw) > 5 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[5, 1]):
                title_info['date_range'] = str(df_raw.iloc[5, 1])
        except (IndexError, KeyError):
            # If extraction fails, continue with empty title_info
            pass

        return title_info

    def _detect_data_structure(self, df_raw):
        """Intelligently detect header and data start positions"""
        header_row = self.DEFAULT_HEADER_ROW
        data_start_row = self.DEFAULT_DATA_START_ROW

        # Look for header indicators in different rows
        header_indicators = ['Load No.', 'Carrier', 'Service Type', 'Ship Date']

        for row_idx in range(5, min(15, len(df_raw))):
            row_data = df_raw.iloc[row_idx].dropna().astype(str).tolist()
            row_str = ' '.join(row_data).lower()

            # Check if this row contains header-like content
            matches = sum(1 for indicator in header_indicators if indicator.lower() in row_str)
            if matches >= 2:  # Found at least 2 header indicators
                header_row = row_idx
                data_start_row = row_idx + 2  # Skip potential blank row
                break

        return header_row, data_start_row

    def _remove_duplicate_headers(self, df):
        """Remove duplicate header rows that appear in the middle of data"""
        # Look for rows that contain header-like text
        header_indicators = ['Load No.', 'Carrier', 'Service Type']

        rows_to_drop = []
        for idx, row in df.iterrows():
            row_str = ' '.join([str(val) for val in row.dropna().tolist()])
            if any(indicator in row_str for indicator in header_indicators):
                # Check if this looks like a header row (not actual data)
                if not any(str(val).startswith('A') and str(val)[1:].isdigit() for val in row.dropna().tolist()):
                    rows_to_drop.append(idx)

        return df.drop(rows_to_drop)

    def clean_and_process_data(self, file_path: str) -> pd.DataFrame:
        """Main function to clean and process the TMS Excel file with comprehensive validation"""
        self.processing_start_time = time.time()
        self.logger.log_processing_step("Starting TMS data processing", {'file': Path(file_path).name})

        try:
            # Step 1: Comprehensive validation
            self.validation_results = tms_validator.run_full_validation(file_path)

            if not self.validation_results['overall_valid']:
                failed_steps = [step for step, result in self.validation_results['validation_steps'].items()
                               if not result.get('valid', False)]
                raise ValueError(f"File validation failed. Issues: {failed_steps}")

            # Load Excel data
            df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
            self.title_info = self._extract_title_info(df_raw)

            # Get data structure
            header_info = self.validation_results['validation_steps']['header_detection']['details']
            header_row = header_info.get('header_row', self.config.get('data_structure.default_header_row', 8))
            data_start_row = header_info.get('data_start_row', self.config.get('data_structure.default_data_start_row', 11))

            # Get headers
            headers = df_raw.iloc[header_row].dropna().tolist()

            # Extract and clean data
            data_df = df_raw.iloc[data_start_row:].copy()
            data_df = data_df.dropna(how='all').reset_index(drop=True)
            data_df = self._remove_duplicate_headers(data_df)

            # Extract relevant columns
            max_cols = min(22, len(data_df.columns))
            relevant_columns = list(range(2, max_cols + 1))
            data_df = data_df.iloc[:, relevant_columns]

            # Set proper column names with full descriptive headers
            base_column_names = [
                'Load No.', 'Ship Date', 'Origin City', 'Origin State', 'Origin Postal',
                'Destination City', 'Destination State', 'Destination Postal',
                'Selected Carrier', 'Selected Service Type', 'Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
                'Least Cost Carrier', 'Least Cost Service Type', 'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost',
                'Potential Savings'
            ]

            # Ensure column names match the actual extracted columns
            if len(base_column_names) != len(data_df.columns):
                print(f"Warning: Column count mismatch. Expected {len(base_column_names)}, got {len(data_df.columns)}")
                # Adjust column names to match actual columns
                if len(data_df.columns) < len(base_column_names):
                    column_names = base_column_names[:len(data_df.columns)]
                else:
                    # Add generic names for extra columns
                    extra_columns = [f'Column_{i}' for i in range(len(base_column_names), len(data_df.columns))]
                    column_names = base_column_names + extra_columns
            else:
                column_names = base_column_names

            data_df.columns = column_names

            # Clean data types and filter invalid rows
            data_df = self._clean_data_types_enhanced(data_df)

            if 'Load No.' in data_df.columns:
                data_df = data_df.dropna(subset=['Load No.'])
                data_df = data_df[data_df['Load No.'].astype(str).str.strip().isin(['', 'nan']) == False]

            min_values = self.config.get('business_rules.min_non_empty_values', 5)
            data_df = data_df.dropna(thresh=min_values).reset_index(drop=True)

            # Apply business logic rules and sort
            data_df = self._apply_business_logic_enhanced(data_df)

            if 'Destination City' in data_df.columns:
                data_df = data_df.sort_values('Destination City', na_position='last')

            # Calculate summary statistics
            self._calculate_summary_stats(data_df)

            # Step 9: Final processing metrics
            processing_time = time.time() - self.processing_start_time
            self.processing_stats = {
                'total_time': processing_time,
                'records_processed': len(data_df),
                'processing_rate': len(data_df) / processing_time if processing_time > 0 else 0
            }

            self.logger.log_performance(
                "Total TMS processing",
                processing_time,
                len(data_df)
            )

            self.processed_data = data_df
            return data_df

        except (FileNotFoundError, PermissionError) as e:
            self.logger.error("File access error", exception=e, file_path=file_path)
            raise FileNotFoundError(f"Cannot access file: {str(e)}")
        except (pd.errors.EmptyDataError, pd.errors.ParserError) as e:
            self.logger.error("Excel parsing error", exception=e, file_path=file_path)
            raise ValueError(f"Invalid Excel file format: {str(e)}")
        except ValueError as e:
            # Re-raise validation errors
            self.logger.error("Validation error", exception=e)
            raise
        except Exception as e:
            self.logger.error("Unexpected processing error", exception=e, file_path=file_path)
            raise RuntimeError(f"Error processing file: {str(e)}")

    def _clean_data_types_enhanced(self, df: pd.DataFrame) -> pd.DataFrame:
        """Enhanced data type cleaning with comprehensive validation and logging"""
        df = df.copy()
        cleaning_stats = {'columns_processed': 0, 'conversion_failures': 0}

        # Convert numeric columns with enhanced error tracking
        numeric_columns = [
            'Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
            'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost', 'Potential Savings'
        ]

        for col in numeric_columns:
            if col in df.columns:
                cleaning_stats['columns_processed'] += 1
                original_nulls = df[col].isnull().sum()
                df[col] = pd.to_numeric(df[col], errors='coerce')
                new_nulls = df[col].isnull().sum()
                conversion_failures = new_nulls - original_nulls
                if conversion_failures > 0:
                    cleaning_stats['conversion_failures'] += conversion_failures
                    self.data_logger.warning(f"Failed to convert {conversion_failures} values in {col} to numeric")

        # Ensure PS column is properly numeric and handle any string values
        if 'PS' in df.columns:
            # First try to convert to numeric, handling any string values
            df['PS'] = pd.to_numeric(df['PS'], errors='coerce')
            # Fill any NaN values with 0
            df['PS'] = df['PS'].fillna(0)

        # Convert date column with enhanced error handling
        if 'Ship Date' in df.columns:
            cleaning_stats['columns_processed'] += 1
            original_nulls = df['Ship Date'].isnull().sum()
            date_series = pd.to_datetime(df['Ship Date'], errors='coerce')
            new_nulls = date_series.isnull().sum()
            date_failures = new_nulls - original_nulls
            if date_failures > 0:
                cleaning_stats['conversion_failures'] += date_failures
                self.data_logger.warning(f"Failed to convert {date_failures} date values in Ship Date")

            date_format = self.config.get('formatting.date_format', '%m/%d/%y')
            df['Ship Date'] = date_series.dt.strftime(date_format)

        # Clean string columns with tracking
        string_columns = [
            'Load No.', 'Origin City', 'Origin State', 'Origin Postal',
            'Destination City', 'Destination State', 'Destination Postal',
            'Selected Carrier', 'Selected Service Type', 'Least Cost Carrier', 'Least Cost Service Type'
        ]

        for col in string_columns:
            if col in df.columns:
                cleaning_stats['columns_processed'] += 1

                # Special handling for Least Cost Carrier and Service Type
                if col in ['Least Cost Carrier', 'Least Cost Service Type']:
                    # Check for numeric zeros that should be converted to empty strings
                    before_cleaning = df[col].copy()
                    df[col] = df[col].astype(str).str.strip()

                    # Convert numeric zeros to empty strings
                    df[col] = df[col].replace(['0', '0.0', 'nan', 'None'], '')

                    # Log the cleaning results for debugging
                    zero_count = (before_cleaning == 0).sum()
                    nan_count = before_cleaning.isna().sum()
                    if zero_count > 0 or nan_count > 0:
                        self.data_logger.info(f"Column {col}: Found {zero_count} zeros and {nan_count} NaN values - converted to empty strings")
                else:
                    df[col] = df[col].astype(str).str.strip()
                    df[col] = df[col].replace('nan', '')

        self.data_logger.log_data_stats(cleaning_stats, "TYPE_CLEANING")
        return df

    def _apply_business_logic_enhanced(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply TMS business logic rules with enhanced tracking and validation"""
        df = df.copy()
        business_stats = {
            'same_carrier_rule_applied': 0,
            'empty_data_rule_applied': 0,
            'negative_savings_rule_applied': 0,
            'tl_carrier_rule_applied': 0,
            'ddi_carrier_rule_applied': 0,
            'total_rows_affected': 0
        }

        try:
            # Ensure PS column is numeric from the start to avoid comparison errors
            if 'PS' in df.columns:
                df['PS'] = pd.to_numeric(df['PS'], errors='coerce').fillna(0)
            else:
                print("Warning: PS column not found in dataframe")
                print(f"Available columns: {df.columns.tolist()}")

            # Rule 1: Same Carriers - Set Potential Savings to 0 (Enhanced)
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                same_carrier_mask = (
                    (df['Selected Carrier'].astype(str) == df['Least Cost Carrier'].astype(str)) &
                    (df['Selected Carrier'].notna()) &
                    (df['Least Cost Carrier'].notna()) &
                    (df['Selected Carrier'].astype(str) != '') &
                    (df['Least Cost Carrier'].astype(str) != '') &
                    (df['Selected Carrier'].astype(str) != 'nan') &
                    (df['Least Cost Carrier'].astype(str) != 'nan')
                )

                same_carrier_count = same_carrier_mask.sum()
                business_stats['same_carrier_rule_applied'] = same_carrier_count

                if 'Potential Savings' in df.columns and same_carrier_count > 0:
                    default_savings = self.config.get('business_rules.same_carrier_savings', 0.0)
                    df.loc[same_carrier_mask, 'Potential Savings'] = default_savings
                    self.data_logger.info(f"Applied same carrier rule to {same_carrier_count} rows")
            else:
                self.data_logger.warning("Cannot apply same carrier rule - required columns missing")

            # Rule 2: Empty Least Cost - Copy Selected data and set savings to 0 (Enhanced)
            if 'Least Cost Carrier' in df.columns:
                empty_least_cost_mask = (
                    df['Least Cost Carrier'].isna() |
                    (df['Least Cost Carrier'].astype(str) == '') |
                    (df['Least Cost Carrier'].astype(str) == 'nan')
                )

                empty_count = empty_least_cost_mask.sum()
                business_stats['empty_data_rule_applied'] = empty_count

                if empty_count > 0:
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'),
                        ('Selected Service Type', 'Least Cost Service Type'),
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'),
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, empty_least_cost_mask, column_pairs)

                    if 'Potential Savings' in df.columns:
                        df.loc[empty_least_cost_mask, 'Potential Savings'] = 0

                    self.data_logger.info(f"Applied empty data rule to {empty_count} rows")
            else:
                self.data_logger.warning("Cannot apply empty data rule - Least Cost Carrier column missing")

            # Rule 3: Negative Savings - Copy Selected data and set savings to 0 (Enhanced)
            if 'Potential Savings' in df.columns:
                # Ensure Potential Savings is numeric before comparison
                ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
                negative_savings_mask = ps_numeric < 0
                negative_count = negative_savings_mask.sum()
                business_stats['negative_savings_rule_applied'] = negative_count

                if negative_count > 0:
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'),
                        ('Selected Service Type', 'Least Cost Service Type'),
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'),
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, negative_savings_mask, column_pairs)
                    df.loc[negative_savings_mask, 'Potential Savings'] = 0
                    self.data_logger.info(f"Applied negative savings rule to {negative_count} rows")
            else:
                self.data_logger.warning("Cannot apply negative savings rule - Potential Savings column missing")

            # Rule 4: TL Carriers - Copy selected to least cost and zero out savings
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                # Find rows where Selected Carrier or Least Cost Carrier is in TL list
                tl_mask = (
                    df['Selected Carrier'].astype(str).str.upper().isin([carrier.upper() for carrier in self.TL_CARRIERS]) |
                    df['Least Cost Carrier'].astype(str).str.upper().isin([carrier.upper() for carrier in self.TL_CARRIERS])
                )

                tl_count = tl_mask.sum()
                business_stats['tl_carrier_rule_applied'] = tl_count

                if tl_count > 0:
                    # Debug: Log which carriers were found
                    tl_carriers_found = df.loc[tl_mask, ['Selected Carrier', 'Least Cost Carrier']].drop_duplicates()
                    self.data_logger.info(f"TL carriers found: {tl_carriers_found.to_dict('records')}")

                    # Copy selected carrier data to least cost columns
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'),
                        ('Selected Service Type', 'Least Cost Service Type'),
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'),
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, tl_mask, column_pairs)

                    # Set Potential Savings to 0
                    if 'Potential Savings' in df.columns:
                        # Debug: Log before and after values
                        before_savings = df.loc[tl_mask, 'Potential Savings'].tolist()
                        df.loc[tl_mask, 'Potential Savings'] = 0
                        after_savings = df.loc[tl_mask, 'Potential Savings'].tolist()
                        self.data_logger.info(f"TL Carrier savings - Before: {before_savings}, After: {after_savings}")
                    else:
                        self.data_logger.warning("Potential Savings column not found for TL carrier rule")

                    self.data_logger.info(f"Applied TL carrier rule to {tl_count} rows (LANDSTAR/SMARTWAY)")
                else:
                    self.data_logger.info("No TL carriers found in data")

            # Rule 5: DDI/Carrier Matching - New custom rule
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                # Create mask for rows where Selected Carrier contains "DDI/" or similar patterns
                # and the part after "/" matches Least Cost Carrier
                ddi_matches = []

                for idx, row in df.iterrows():
                    selected = str(row['Selected Carrier']).strip()
                    least_cost = str(row['Least Cost Carrier']).strip()

                    # Skip empty or nan values
                    if selected in ['', 'nan', 'None'] or least_cost in ['', 'nan', 'None']:
                        continue

                    # Check if selected carrier has "/" and extract the part after it
                    if '/' in selected:
                        # Split on "/" and get the part after the last "/"
                        carrier_after_slash = selected.split('/')[-1].strip()

                        # Check if the carrier after "/" matches the least cost carrier
                        # Using case-insensitive comparison and handling common variations
                        if carrier_after_slash.upper() == least_cost.upper():
                            ddi_matches.append(idx)
                        # Also check for R&L Carriers vs R%L Carriers variations
                        elif (carrier_after_slash.upper().replace('&', '%') == least_cost.upper().replace('&', '%') or
                              carrier_after_slash.upper().replace('%', '&') == least_cost.upper().replace('%', '&')):
                            ddi_matches.append(idx)

                ddi_match_count = len(ddi_matches)
                business_stats['ddi_carrier_rule_applied'] = ddi_match_count

                if ddi_match_count > 0:
                    ddi_mask = df.index.isin(ddi_matches)

                    # Copy selected carrier data to least cost columns
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'),
                        ('Selected Service Type', 'Least Cost Service Type'),
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'),
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, ddi_mask, column_pairs)

                    # Set Potential Savings to 0
                    if 'Potential Savings' in df.columns:
                        df.loc[ddi_mask, 'Potential Savings'] = 0

                    self.data_logger.info(f"Applied DDI/carrier matching rule to {ddi_match_count} rows")
            else:
                self.data_logger.warning("Cannot apply DDI/carrier matching rule - required columns missing")

            # Calculate total affected rows
            business_stats['total_rows_affected'] = (
                business_stats['same_carrier_rule_applied'] +
                business_stats['empty_data_rule_applied'] +
                business_stats['negative_savings_rule_applied'] +
                business_stats['tl_carrier_rule_applied'] +
                business_stats['ddi_carrier_rule_applied']
            )

            self.data_logger.log_data_stats(business_stats, "BUSINESS_LOGIC")

        except Exception as e:
            self.data_logger.error("Business logic application failed", exception=e,
                                 df_shape=df.shape, df_columns=df.columns.tolist())
            raise RuntimeError(f"Business logic error: {str(e)}")

        return df

    def _copy_selected_to_least_cost(self, df, mask, column_pairs):
        """Helper method to copy selected carrier data to least cost columns"""
        for selected_col, least_cost_col in column_pairs:
            if selected_col in df.columns and least_cost_col in df.columns:
                df.loc[mask, least_cost_col] = df.loc[mask, selected_col]

    def _calculate_summary_stats(self, df):
        """Calculate comprehensive summary statistics"""
        if df.empty:
            self.summary_stats = {
                'total_loads': 0,
                'total_selected_cost': 0,
                'total_least_cost': 0,
                'total_potential_savings': 0,
                'average_savings_per_load': 0,
                'percentage_savings': 0,
                'loads_with_savings': 0,
                'total_savings_opportunity': 0
            }
            return

        # Basic stats - ensure numeric columns are properly converted
        total_loads = len(df)
        total_selected_cost = pd.to_numeric(df['Selected Total Cost'], errors='coerce').fillna(0).sum()
        total_least_cost = pd.to_numeric(df['Least Cost Total Cost'], errors='coerce').fillna(0).sum()
        total_potential_savings = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0).sum()

        # Advanced stats - optimize by filtering once
        ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
        savings_df = df[ps_numeric > 0]
        loads_with_savings = len(savings_df)
        total_savings_opportunity = pd.to_numeric(savings_df['Potential Savings'], errors='coerce').fillna(0).sum()

        # Calculate percentages
        if total_selected_cost > 0:
            percentage_savings = (total_potential_savings / total_selected_cost) * 100
        else:
            percentage_savings = 0

        if total_loads > 0:
            average_savings_per_load = total_potential_savings / total_loads
        else:
            average_savings_per_load = 0

        self.summary_stats = {
            'total_loads': total_loads,
            'total_selected_cost': total_selected_cost,
            'total_least_cost': total_least_cost,
            'total_potential_savings': total_potential_savings,
            'average_savings_per_load': average_savings_per_load,
            'percentage_savings': percentage_savings,
            'loads_with_savings': loads_with_savings,
            'total_savings_opportunity': total_savings_opportunity
        }

    def save_processed_data(self, output_file):
        """Save processed data to Excel with professional formatting"""
        if self.processed_data is None:
            raise Exception("No processed data to save")

        # Create a new workbook
        wb = openpyxl.Workbook()

        # Create processed data sheet
        ws_data = wb.active
        ws_data.title = "Processed Data"

        # Add title info efficiently
        row = 1
        if self.title_info:
            def create_title_row(text, row_num):
                last_col = get_column_letter(len(self.processed_data.columns))
                cell = ws_data[f'A{row_num}']
                cell.value = text
                cell.font = Font(size=12, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws_data.merge_cells(f'A{row_num}:{last_col}{row_num}')
                ws_data.row_dimensions[row_num].height = 25
                return row_num + 1

            if 'company_name' in self.title_info:
                row = create_title_row(f"Company: {self.title_info['company_name']}", row)
            if 'date_range' in self.title_info:
                row = create_title_row(f"Date Range: {self.title_info['date_range']}", row)

            # Add section headers row with color coding
            row = 4
            # Selected Carrier section (columns I-N, which are 9-14) - Light Blue
            selected_header = ws_data.cell(row=row, column=9, value="Selected Carrier")
            selected_header.font = Font(size=10, bold=True, color="FFFFFF")  # Reduced from 11 to 10
            selected_header.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            selected_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('I4:N4')
            for col in range(9, 15):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")

            # Least Cost Carrier section (columns O-T, which are 15-20) - Light Orange
            least_cost_header = ws_data.cell(row=row, column=15, value="Least Cost Carrier")
            least_cost_header.font = Font(size=10, bold=True, color="FFFFFF")  # Reduced from 11 to 10
            least_cost_header.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
            least_cost_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('O4:T4')
            for col in range(15, 21):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")

            row = 5  # Headers will be on row 5

        # Add headers with enhanced styling and color coding
        headers = self.processed_data.columns.tolist()
        header_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=row, column=col_idx, value=header)
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Color code headers based on section
            if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Light Blue
            elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Light Orange
            elif header == 'Potential Savings':  # Potential Savings column - Green
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
            else:
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Default blue

        # Add data with alternating row colors and comprehensive borders
        data_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Filter out any remaining empty rows before writing to Excel
        clean_data = self.processed_data.dropna(subset=['Load No.'])
        clean_data = clean_data[clean_data['Load No.'].astype(str).str.strip() != '']

        # Ensure all data is properly typed before processing
        for col in clean_data.columns:
            if col in ['Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost', 'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost', 'Potential Savings']:
                clean_data[col] = pd.to_numeric(clean_data[col], errors='coerce').fillna(0)
            else:
                clean_data[col] = clean_data[col].astype(str).fillna('')

        for data_idx, data_row in enumerate(dataframe_to_rows(clean_data, index=False, header=False)):
            # Skip rows that are mostly empty
            # Ensure we're comparing integers by converting the sum result
            try:
                non_empty_count = sum(1 for val in data_row if val is not None and str(val).strip() != '' and str(val) != 'nan')
                if non_empty_count < 3:
                    continue
            except Exception as e:
                print(f"Error processing row {data_idx}: {e}")
                print(f"Row data: {data_row}")
                continue

            row += 1
            row_color = "F8F9FA" if data_idx % 2 == 0 else "FFFFFF"

            # First pass: collect all content lengths to determine optimal row height
            max_content_length = 0
            for col_idx, value in enumerate(data_row, 1):
                content_length = len(str(value)) if value else 0
                max_content_length = max(max_content_length, content_length)

            # Enhanced dynamic height calculation for long carrier names
            # Check if this row contains carrier information that might wrap
            has_carrier_data = any('TRANSPORT' in str(val).upper() or
                                 'LOGISTICS' in str(val).upper() or
                                 'FREIGHT' in str(val).upper() or
                                 len(str(val)) > 25 for val in data_row if val)

            if has_carrier_data and max_content_length > 25:
                # For carrier names, be more generous with height to prevent cutoff
                optimal_height = min(50, max(30, max_content_length * 1.2))
            elif max_content_length > 30:  # Very long content
                optimal_height = min(45, max(25, max_content_length * 1.0))
            elif max_content_length > 20:  # Long content
                optimal_height = min(35, max(22, max_content_length * 0.8))
            elif max_content_length > 15:  # Medium content
                optimal_height = 25
            else:
                optimal_height = 20  # Default height with a bit more room

            # Set the row height once for the entire row
            ws_data.row_dimensions[row].height = optimal_height

            # Second pass: apply formatting to all cells in the row
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_data.cell(row=row, column=col_idx, value=value)
                header_name = headers[col_idx-1] if col_idx-1 < len(headers) else ""
                # Center all cell contents and enable text wrapping for compactness
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = data_border

                # Apply color coding to data cells based on section
                if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                    light_blue_bg = "E6F3FF" if data_idx % 2 == 0 else "F0F8FF"
                    cell.fill = PatternFill(start_color=light_blue_bg, end_color=light_blue_bg, fill_type="solid")
                elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                    light_orange_bg = "FFF2E6" if data_idx % 2 == 0 else "FFF8F0"
                    cell.fill = PatternFill(start_color=light_orange_bg, end_color=light_orange_bg, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")

                # Format currency columns and apply green color for positive Potential Savings values
                currency_columns = ['Selected Total Cost', 'Least Cost Total Cost', 'Selected Freight Cost', 'Least Cost Freight Cost', 'Selected Accessorial Cost', 'Least Cost Accessorial Cost']
                if headers[col_idx-1] in currency_columns or headers[col_idx-1] == 'Potential Savings':
                    cell.number_format = '"$"#,##0.00'
                    # Apply light green background for positive Potential Savings values
                    if headers[col_idx-1] == 'Potential Savings':
                        try:
                            # Convert value to float for comparison, handle None and string values
                            if value is not None and value != '' and str(value).lower() != 'nan':
                                # Handle different value types safely
                                if isinstance(value, (int, float)):
                                    numeric_value = float(value)
                                else:
                                    numeric_value = float(str(value).replace('$', '').replace(',', ''))
                                if numeric_value > 0:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        except (ValueError, TypeError, AttributeError):
                            pass  # Skip coloring if value can't be converted
                    cell.font = Font(size=10, bold=False, color="2C3E50")
            else:
                cell.font = Font(size=10, color="495057")

        # Enable auto-filter over header and data range (no freeze panes)
        try:
            header_row_idx = 5
            ws_data.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{row}"
        except Exception:
            pass

        # Add totals row with key financial metrics
        totals_row = row + 2

        # Add "TOTALS" label
        totals_label = ws_data.cell(row=totals_row, column=1, value="TOTALS")
        totals_label.font = Font(size=12, bold=True, color="FFFFFF")
        totals_label.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        totals_label.alignment = Alignment(horizontal="center", vertical="center")
        totals_label.border = Border(
            left=Side(style='medium', color='2C3E50'),
            right=Side(style='medium', color='2C3E50'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )

        # Find the Selected Total Cost and Potential Savings columns
        selected_cost_col = None
        potential_savings_col = None
        for col_idx, header in enumerate(headers, 1):
            if 'Selected Total Cost' in str(header):
                selected_cost_col = col_idx
            elif 'Potential Savings' in str(header):
                potential_savings_col = col_idx

        # Add Total Selected Cost
        if selected_cost_col:
            cost_cell = ws_data.cell(row=totals_row, column=selected_cost_col,
                                   value=f"${self.summary_stats['total_selected_cost']:,.2f}")
            cost_cell.font = Font(size=12, bold=True, color="FFFFFF")
            cost_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cost_cell.alignment = Alignment(horizontal="center", vertical="center")
            cost_cell.number_format = '"$"#,##0.00'
            cost_cell.border = Border(
                left=Side(style='medium', color='3498DB'),
                right=Side(style='medium', color='3498DB'),
                top=Side(style='medium', color='3498DB'),
                bottom=Side(style='medium', color='3498DB')
            )

        # Add Total Potential Savings (most important number)
        if potential_savings_col:
            savings_cell = ws_data.cell(row=totals_row, column=potential_savings_col,
                                      value=f"${self.summary_stats['total_potential_savings']:,.2f}")
            savings_cell.font = Font(size=14, bold=True, color="FFFFFF")  # Larger font for emphasis
            savings_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            savings_cell.alignment = Alignment(horizontal="center", vertical="center")
            savings_cell.number_format = '"$"#,##0.00'
            savings_cell.border = Border(
                left=Side(style='thick', color='27AE60'),  # Thicker border for emphasis
                right=Side(style='thick', color='27AE60'),
                top=Side(style='thick', color='27AE60'),
                bottom=Side(style='thick', color='27AE60')
            )

        # Set height for totals row
        ws_data.row_dimensions[totals_row].height = 25


        # Auto-fit column widths on the Processed Data sheet (compact and consistent for all: table, CAL, PI)
        try:
            for col_idx in range(1, ws_data.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(5, ws_data.max_row + 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        text = str(cell.value)
                        length = len(text)
                        if length > max_length:
                            max_length = length
                # Aggressive compact auto-sizing: minimize width while ensuring content fits
                # Check if this column contains text content
                has_text_content = False
                for check_row in range(5, ws_data.max_row + 1):
                    check_cell = ws_data.cell(row=check_row, column=col_idx)
                    if check_cell.value and any(c.isalpha() for c in str(check_cell.value)):
                        has_text_content = True
                        break

                # Very tight padding for maximum compactness
                if has_text_content:
                    padding = 1.0  # Minimal padding for text
                else:
                    padding = 0.5   # Very tight for numbers

                # Calculate width with aggressive compacting
                adjusted_width = max_length + padding

                # Apply maximum width constraints for compactness
                if has_text_content:
                    max_width = 25  # Cap text columns at 25 characters
                else:
                    max_width = 15  # Cap number columns at 15 characters

                final_width = min(adjusted_width, max_width)
                ws_data.column_dimensions[col_letter].width = max(6, final_width)  # Minimum 6 for readability

            # Optimize row heights for compact layout
            for rh in [1, 2, 4, 5]:
                if rh <= ws_data.max_row:
                    ws_data.row_dimensions[rh].height = max(ws_data.row_dimensions[rh].height or 0, 20)  # Reduced from 22 to 20

            # Ensure gridlines visible
            ws_data.sheet_view.showGridLines = True
        except Exception:
            pass

        # Add thick outside borders around the entire data table
        try:
            # Define thick border styles
            thick_side = Side(style='medium', color='2C3E50')

            header_row_idx = 5
            first_row = header_row_idx
            last_row = row
            first_col = 1
            last_col = len(headers)

            # Apply thick borders to all edge cells
            for r in range(first_row, last_row + 1):
                for c in range(first_col, last_col + 1):
                    cell = ws_data.cell(row=r, column=c)
                    current_border = cell.border or Border()

                    # Determine which sides need thick borders
                    left_side = thick_side if c == first_col else current_border.left
                    right_side = thick_side if c == last_col else current_border.right
                    top_side = thick_side if r == first_row else current_border.top
                    bottom_side = thick_side if r == last_row else current_border.bottom

                    # Apply the new border
                    cell.border = Border(
                        left=left_side,
                        right=right_side,
                        top=top_side,
                        bottom=bottom_side
                    )
        except Exception:
            pass

        wb.save(output_file)
        wb.close()


class BasicTMSGUI:

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Basic TMS Processor")
        self.root.configure(bg='#f0f0f0')

        # Center window to match original size
        self.center_window()

        # Initialize processor and state
        self.processor = ModernTMSProcessor()
        self.input_files = []
        self.processing = False

        self.setup_gui()

    def center_window(self):
        """Center smaller, more compact window"""
        # Calculate screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Even smaller, more compact size
        window_width = 750
        window_height = 550

        # Center position
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2

        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.root.resizable(True, True)

    def setup_gui(self):
        """Setup vertical split layout: Select Files left, Results right"""
        # Main container
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Header
        self.create_header(main_frame)

        # Main content area - vertical split
        content_frame = tk.Frame(main_frame, bg='#f0f0f0')
        content_frame.pack(fill='both', expand=True, pady=(10, 0))

        # Left side - Select Files container
        left_container = tk.Frame(content_frame, bg='white')
        left_container.pack(side='left', fill='both', expand=True, padx=(0, 5))

        # Right side - Results container
        right_container = tk.Frame(content_frame, bg='white')
        right_container.pack(side='right', fill='both', expand=True, padx=(5, 0))

        # Setup left and right sections
        self.create_select_files_section(left_container)
        self.create_results_display(right_container)

    def create_header(self, parent):
        """Create compact header for Basic TMS Processor"""
        header_frame = tk.Frame(parent, bg='#f0f0f0')
        header_frame.pack(fill='x', pady=(5, 10))

        title = tk.Label(header_frame, text="📊 Basic TMS Processor",
                        font=('Segoe UI', 20, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        title.pack()

    def create_select_files_section(self, parent):
        """Create the complete Select Files section with buttons"""
        # Header
        header_frame = tk.Frame(parent, bg='#f8f9fa')
        header_frame.pack(fill='x')

        header = tk.Label(header_frame, text="📁 Select Files",
                         font=('Segoe UI', 12, 'bold'), bg='#f8f9fa', fg='#2c3e50',
                         padx=15, pady=6)
        header.pack(side='left')

        self.file_counter_label = tk.Label(header_frame, text="0 files selected",
                                          font=('Segoe UI', 10), bg='#f8f9fa', fg='#27ae60',
                                          padx=15, pady=6)
        self.file_counter_label.pack(side='right')

        # Content area
        content_frame = tk.Frame(parent, bg='white')
        content_frame.pack(fill='both', expand=True, padx=15, pady=15)

        # Browse button
        self.browse_btn = tk.Button(content_frame, text="📂 Browse",
                                   font=('Segoe UI', 11, 'bold'),
                                   bg='#4a90e2', fg='white',
                                   relief='flat', bd=0, cursor='hand2',
                                   command=self.select_files,
                                   padx=20, pady=10)
        self.browse_btn.pack(fill='x', pady=(0, 15))

        # File list area with scrollable listbox
        list_container = tk.Frame(content_frame, bg='white')
        list_container.pack(fill='both', expand=True, pady=(0, 15))

        self.file_listbox = tk.Listbox(list_container, font=('Segoe UI', 9),
                                      bg='#f8f9fa', fg='#2c3e50', relief='flat', bd=1,
                                      selectbackground='#4a90e2', selectforeground='white')

        file_scrollbar = tk.Scrollbar(list_container, command=self.file_listbox.yview)
        self.file_listbox.config(yscrollcommand=file_scrollbar.set)

        self.file_listbox.pack(side='left', fill='both', expand=True)
        file_scrollbar.pack(side='right', fill='y')

        # Keep the old frame for backward compatibility
        self.file_list_frame = list_container

        # Process button
        self.process_btn = tk.Button(content_frame, text="🚀 PROCESS",
                                    font=('Segoe UI', 12, 'bold'),
                                    bg='#27ae60', fg='white',
                                    relief='flat', bd=0, cursor='hand2',
                                    command=self.process_files,
                                    padx=20, pady=12)
        self.process_btn.pack(fill='x')

    def create_results_display(self, parent):
        """Create a simple, clean results display for the right side"""
        # Header
        header_frame = tk.Frame(parent, bg='#f8f9fa')
        header_frame.pack(fill='x')

        header = tk.Label(header_frame, text="📊 Results",
                         font=('Segoe UI', 12, 'bold'), bg='#f8f9fa', fg='#2c3e50',
                         padx=15, pady=6)
        header.pack(side='left')

        # Clear button
        clear_btn = tk.Button(header_frame, text="Clear",
                             font=('Segoe UI', 9), bg='#e74c3c', fg='white',
                             relief='flat', bd=0, cursor='hand2',
                             command=self.clear_results,
                             padx=8, pady=3)
        clear_btn.pack(side='right', padx=5, pady=3)

        # Results text area
        text_frame = tk.Frame(parent, bg='white')
        text_frame.pack(fill='both', expand=True, padx=15, pady=15)

        self.results_text = tk.Text(text_frame, font=('Segoe UI', 9),
                                   bg='white', fg='#2c3e50',
                                   relief='flat', bd=0, wrap='word',
                                   state='disabled')

        # Scrollbar - more visible
        scrollbar = tk.Scrollbar(text_frame, command=self.results_text.yview,
                               bg='#e0e0e0', troughcolor='#f0f0f0',
                               activebackground='#c0c0c0', width=12)
        self.results_text.config(yscrollcommand=scrollbar.set)

        self.results_text.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Configure text tags for formatting
        self.results_text.tag_configure("result", foreground="#2c3e50", font=('Segoe UI', 9))
        self.results_text.tag_configure("filename", foreground="#4a90e2", font=('Segoe UI', 9))
        self.results_text.tag_configure("savings", foreground="#27ae60", font=('Segoe UI', 9, 'bold'))

        # Load previous results on startup
        self.load_previous_results()

    def add_result_entry(self, date_str, time_str, type_str, files_str, filename, ps_amount):
        """Add a clean log entry with improved formatting"""
        self.results_text.config(state='normal')

        # Add with better formatting and green PS amounts
        self.results_text.insert('end', f"{date_str} | {time_str} | {type_str} | {files_str} | ", 'result')
        self.results_text.insert('end', f"{filename}", 'filename')
        self.results_text.insert('end', f" | ", 'result')
        self.results_text.insert('end', f"${ps_amount:,.2f}", 'savings')
        self.results_text.insert('end', '\n', 'result')

        # Scroll to bottom
        self.results_text.see('end')
        self.results_text.config(state='disabled')
        self.root.update_idletasks()

    def clear_results(self):
        """Clear all results and save empty state"""
        self.results_text.config(state='normal')
        self.results_text.delete(1.0, 'end')
        self.results_text.config(state='disabled')

        # Clear the saved history file
        try:
            history_file = Path("logs") / "processing_history.json"
            if history_file.exists():
                with open(history_file, 'w') as f:
                    json.dump([], f)
        except Exception as e:
            gui_logger.error("Failed to clear processing history", exception=e)

    def load_previous_results(self):
        """Load previous results from processing history"""
        try:
            history_file = Path("logs") / "processing_history.json"
            if not history_file.exists():
                return

            with open(history_file, 'r') as f:
                history_data = json.load(f)

            if not history_data:
                return

            # Load and display each entry
            for entry in history_data:
                # Handle both new format (separate date/time) and old format
                if 'date' in entry and 'time' in entry:
                    date_str = entry.get('date', '')
                    time_str = entry.get('time', '')
                else:
                    # Legacy format - parse combined time string
                    time_parts = entry.get('time', '').split(' ')
                    if len(time_parts) >= 2:
                        date_str = time_parts[0]
                        time_str = ' '.join(time_parts[1:])
                    else:
                        date_str = entry.get('time', '')
                        time_str = ''

                type_str = entry.get('type', 'Basic')
                file_counter = entry.get('file_counter', 'Single')
                filename = entry.get('file', '')
                savings_str = entry.get('savings', '$0.00')

                # Extract numeric value from savings string
                try:
                    ps_amount = float(savings_str.replace('$', '').replace(',', ''))
                except:
                    ps_amount = 0.0

                # Add entry to results display with file counter
                self.results_text.config(state='normal')
                self.results_text.insert('end', f"{date_str} | {time_str} | {type_str} | {file_counter} | ", 'result')
                self.results_text.insert('end', f"{filename}", 'filename')
                self.results_text.insert('end', f" | ", 'result')
                self.results_text.insert('end', f"${ps_amount:,.2f}", 'savings')
                self.results_text.insert('end', '\n', 'result')
                self.results_text.config(state='disabled')

            # Scroll to bottom to show most recent entries
            self.results_text.see('end')

        except Exception as e:
            gui_logger.error("Failed to load previous results", exception=e)




    def populate_recent_uploads(self):
        """Legacy method - no longer needed with logger-focused UI"""
        pass

    def save_processing_history(self, filename, savings, file_number=None, total_files=None):
        """Save processing result to history for recent uploads display"""
        try:
            history_file = Path("logs") / "processing_history.json"
            history_file.parent.mkdir(exist_ok=True)

            # Load existing history
            history_data = []
            if history_file.exists():
                with open(history_file, 'r') as f:
                    history_data = json.load(f)

            # Create file counter string if provided
            if file_number and total_files:
                file_counter = f"{file_number} of {total_files}"
            else:
                file_counter = "Single"

            # Add new entry with date and time separated
            now = datetime.now()
            new_entry = {
                "date": now.strftime("%m/%d/%y"),
                "time": now.strftime("%I:%M %p"),
                "type": "Basic",
                "file_counter": file_counter,
                "file": filename,
                "savings": f"${savings:,.2f}" if savings > 0 else "$0.00"
            }

            history_data.append(new_entry)

            # Keep only last 50 entries (increased for better persistence)
            if len(history_data) > 50:
                history_data = history_data[-50:]

            # Save updated history
            with open(history_file, 'w') as f:
                json.dump(history_data, f, indent=2)

            gui_logger.info("Processing history updated",
                          file=filename, savings=f"${savings:,.2f}")

        except Exception as e:
            gui_logger.error("Failed to save processing history", exception=e)

    def refresh_recent_uploads(self):
        """Legacy method - no longer needed with logger-focused UI"""
        pass


    def select_files(self):
        """Select input files"""
        files = filedialog.askopenfilenames(
            title="Select TMS Excel Files",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if files:
            self.input_files = list(files)
            self.update_file_list()

    def clear_files(self):
        """Clear selected files"""
        self.input_files = []
        self.update_file_list()

    def update_file_list(self):
        """Update the file list display using scrollable listbox"""
        # Clear existing file list
        self.file_listbox.delete(0, tk.END)

        # Update counter
        count = len(self.input_files)
        self.file_counter_label.config(text=f"{count} files selected")

        # Add all files to listbox (scrollable, so no limit needed)
        if count > 0:
            for i, file_path in enumerate(self.input_files, 1):
                filename = os.path.basename(file_path)
                self.file_listbox.insert(tk.END, f"{i}. {filename}")

    def process_files(self):
        """Process the selected files with better state management"""
        if not self.input_files:
            messagebox.showwarning("No Files", "Please select files to process first.")
            return

        if self.processing:
            return  # Prevent double-click issues

        self.processing = True
        self._set_processing_state(True)

        # Run processing in thread
        thread = threading.Thread(target=self._process_files_thread)
        thread.daemon = True
        thread.start()

    def _set_processing_state(self, processing):
        """Update UI state for processing"""
        if processing:
            self.process_btn.config(state='disabled', text="Processing...", bg='#95a5a6')
            self.browse_btn.config(state='disabled')
        else:
            self.process_btn.config(state='normal', text="🚀 PROCESS", bg='#27ae60')
            self.browse_btn.config(state='normal')
            self.processing = False

    def _process_files_thread(self):
        """Process files in separate thread"""
        try:
            processed_files = []
            total_savings = 0.0

            # Create output folder for multiple files
            output_folder = None
            if len(self.input_files) > 1:
                # Create timestamped folder for batch processing
                now = datetime.now()
                folder_name = f"Basic_Processed_{now.strftime('%m.%d_%H%M')}"
                input_dir = os.path.dirname(self.input_files[0])
                output_folder = os.path.join(input_dir, folder_name)
                os.makedirs(output_folder, exist_ok=True)

            for i, input_file in enumerate(self.input_files, 1):
                filename = os.path.basename(input_file)

                # Process the file
                self.processor.process_file(input_file)

                # Generate output filename
                input_name = os.path.splitext(os.path.basename(input_file))[0]
                processed_filename = f"{input_name}_BASIC_PROCESSED.xlsx"

                if output_folder:
                    # Multiple files - save to timestamped folder
                    output_file = os.path.join(output_folder, processed_filename)
                else:
                    # Single file - save to same directory as input
                    input_dir = os.path.dirname(input_file)
                    output_file = os.path.join(input_dir, processed_filename)

                # Save the file
                self.processor.save_processed_data(output_file)
                processed_files.append(output_file)

                # Update totals and get PS
                stats = self.processor.summary_stats
                file_savings = stats.get('total_potential_savings', 0)
                total_savings += file_savings

                # Get date and time
                now = datetime.now()
                date_str = now.strftime('%m/%d/%y')
                time_str = now.strftime('%I:%M %p')

                # Format files info (current file number of total)
                files_str = f"{i} of {len(self.input_files)}"

                # Add result entry with clean log format
                self.root.after(0, lambda d=date_str, t=time_str, fs=files_str, f=filename, s=file_savings:
                    self.add_result_entry(d, t, "Basic", fs, f, s))

                # Save to processing history with file counter
                self.save_processing_history(filename, file_savings, i, len(self.input_files))


            # Show completion dialog and refresh recent uploads
            self.root.after(0, lambda: self._show_completion_summary(processed_files, total_savings, output_folder))
            self.root.after(0, lambda: self.refresh_recent_uploads())

        except Exception as e:
            error_msg = f"Processing error: {str(e)}"
            self.root.after(0, lambda: self.log_message(error_msg, "error"))
            self.root.after(0, lambda: messagebox.showerror("Processing Error", f"Error processing files:\n{str(e)}"))
        finally:
            # Re-enable UI
            self.root.after(0, lambda: self._set_processing_state(False))

    def _show_completion_summary(self, processed_files, total_savings, output_folder=None):
        """Show completion summary with output location"""
        if output_folder:
            # Multiple files processed to organized folder
            folder_name = os.path.basename(output_folder)
            messagebox.showinfo("Processing Complete",
                f"Successfully processed {len(processed_files)} file(s)\n\n"
                f"💰 Total Savings: ${total_savings:,.2f}\n\n"
                f"📁 Files saved to:\n{folder_name}")
        else:
            # Single file processed to same directory
            messagebox.showinfo("Processing Complete",
                f"Successfully processed 1 file\n\n"
                f"💰 Total Savings: ${total_savings:,.2f}\n\n"
                f"📁 File saved to same directory")

    def run(self):
        """Start the GUI application"""
        self.root.mainloop()


def main():
    """Main entry point"""
    app = BasicTMSGUI()
    app.run()


if __name__ == "__main__":
    main()
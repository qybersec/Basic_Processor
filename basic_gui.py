#!/usr/bin/env python3
"""
Basic TMS Processor GUI - Simplified GUI for Basic Report Processing
Extracted from BVC_Automator focusing only on Basic report functionality.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import time
from datetime import datetime


class BasicTMSProcessor:
    """Core TMS processor with proper PS calculation"""

    def __init__(self):
        # Default file structure settings
        self.DEFAULT_HEADER_ROW = 7  # Row 8 in Excel (0-indexed)
        self.DEFAULT_DATA_START_ROW = 10  # Row 11 in Excel (0-indexed)

        # TL carriers that require special processing
        self.TL_CARRIERS = {
            'LANDSTAR RANGER INC',
            'SMARTWAY TRANSPORTATION INC',
            'SMARTWAY CORPORATION INC'
        }

        # Expected column mapping (handles PS -> Potential Savings)
        self.COLUMN_MAPPING = {
            'PS': 'Potential Savings'
        }

        # Storage for processed data
        self.processed_data = None
        self.title_info = {}
        self.summary_stats = {}

    def process_file(self, file_path):
        """Main processing method with proper PS handling"""
        print(f"🔄 Processing: {os.path.basename(file_path)}")

        try:
            # Read Excel file
            df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
            print(f"📄 File loaded: {len(df_raw)} total rows")

            # Extract title information
            self.title_info = self._extract_title_info(df_raw)

            # Extract headers and data
            headers = df_raw.iloc[self.DEFAULT_HEADER_ROW].fillna('').astype(str).tolist()
            data_df = df_raw.iloc[self.DEFAULT_DATA_START_ROW:].copy()
            data_df.columns = headers

            print(f"📊 Original columns: {list(data_df.columns)}")

            # Handle column mapping (PS -> Potential Savings)
            for old_col, new_col in self.COLUMN_MAPPING.items():
                if old_col in data_df.columns:
                    data_df = data_df.rename(columns={old_col: new_col})
                    print(f"🔄 Mapped column: {old_col} -> {new_col}")

            # Clean the data
            data_df = self._clean_data(data_df)
            print(f"🧹 Data cleaned: {len(data_df)} valid records")

            # Calculate PS if not present but cost columns exist
            data_df = self._calculate_potential_savings(data_df)

            # Apply business logic rules
            data_df = self._apply_business_rules(data_df)

            # Calculate summary statistics
            self._calculate_summary_stats(data_df)

            # Sort by destination
            if 'Destination City' in data_df.columns:
                data_df = data_df.sort_values('Destination City', na_position='last')

            self.processed_data = data_df
            return data_df

        except Exception as e:
            print(f"❌ Error processing file: {str(e)}")
            raise

    def _calculate_potential_savings(self, df):
        """Calculate Potential Savings if missing"""
        df = df.copy()

        # If PS already exists and has valid data, keep it
        if 'Potential Savings' in df.columns:
            ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce')
            if ps_numeric.notna().sum() > 0:
                print("💰 Using existing Potential Savings column")
                return df

        # Calculate PS from cost difference
        if 'Selected Total Cost' in df.columns and 'Least Cost Total Cost' in df.columns:
            selected_cost = pd.to_numeric(df['Selected Total Cost'], errors='coerce').fillna(0)
            least_cost = pd.to_numeric(df['Least Cost Total Cost'], errors='coerce').fillna(0)

            # Calculate potential savings (Selected - Least Cost)
            df['Potential Savings'] = selected_cost - least_cost

            # Handle cases where least cost is 0 or missing
            mask_zero_least = least_cost == 0
            df.loc[mask_zero_least, 'Potential Savings'] = 0

            calculated_savings = (df['Potential Savings'] > 0).sum()
            print(f"💰 Calculated Potential Savings for {len(df)} rows ({calculated_savings} with positive savings)")
        else:
            # Create empty PS column if cost columns missing
            df['Potential Savings'] = 0
            print("⚠️ Warning: Cost columns missing, created empty Potential Savings column")

        return df

    def _extract_title_info(self, df_raw):
        """Extract company name and date range from top rows"""
        title_info = {}

        try:
            # Extract company name (typically row 4, column B)
            if len(df_raw) > 3 and len(df_raw.columns) > 1:
                company = df_raw.iloc[3, 1]
                if pd.notna(company):
                    title_info['company_name'] = str(company)

            # Extract date range (typically row 6, column B)
            if len(df_raw) > 5 and len(df_raw.columns) > 1:
                date_range = df_raw.iloc[5, 1]
                if pd.notna(date_range):
                    title_info['date_range'] = str(date_range)

        except Exception as e:
            print(f"⚠️ Could not extract title info: {e}")

        return title_info

    def _clean_data(self, df):
        """Clean and validate the data"""
        df = df.copy()

        # Remove rows where Load No. is missing
        if 'Load No.' in df.columns:
            initial_count = len(df)
            df = df.dropna(subset=['Load No.'])
            df = df[df['Load No.'].astype(str).str.strip() != '']
            df = df[df['Load No.'].astype(str) != 'nan']
            removed = initial_count - len(df)
            if removed > 0:
                print(f"🗑️ Removed {removed} rows with missing Load No.")

        # Clean numeric columns
        numeric_columns = [
            'Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
            'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost',
            'Potential Savings'
        ]

        for col in numeric_columns:
            if col in df.columns:
                original_values = df[col].copy()
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Clean string columns
        string_columns = [
            'Selected Carrier', 'Selected Service Type', 'Least Cost Carrier', 'Least Cost Service Type'
        ]

        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')

        # Remove mostly empty rows
        initial_count = len(df)
        df = df.dropna(thresh=5)
        df = df.reset_index(drop=True)
        removed = initial_count - len(df)
        if removed > 0:
            print(f"🗑️ Removed {removed} mostly empty rows")

        return df

    def _apply_business_rules(self, df):
        """Apply the core TMS business logic rules with detailed logging"""
        df = df.copy()
        rules_applied = []

        print("📋 Applying business rules...")

        # Rule 1: Same Carrier Rule
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            same_carrier_mask = (
                (df['Selected Carrier'].astype(str) == df['Least Cost Carrier'].astype(str)) &
                (df['Selected Carrier'].notna()) &
                (df['Least Cost Carrier'].notna()) &
                (df['Selected Carrier'].astype(str) != '') &
                (df['Least Cost Carrier'].astype(str) != '')
            )

            count = same_carrier_mask.sum()
            if count > 0 and 'Potential Savings' in df.columns:
                df.loc[same_carrier_mask, 'Potential Savings'] = 0
                rules_applied.append(f"Same Carrier Rule: {count} rows")

        # Rule 2: Empty Least Cost Data Rule
        if 'Least Cost Carrier' in df.columns:
            empty_mask = (
                df['Least Cost Carrier'].isna() |
                (df['Least Cost Carrier'].astype(str) == '') |
                (df['Least Cost Carrier'].astype(str) == 'nan')
            )

            count = empty_mask.sum()
            if count > 0:
                # Copy selected data to least cost
                column_pairs = [
                    ('Selected Carrier', 'Least Cost Carrier'),
                    ('Selected Service Type', 'Least Cost Service Type'),
                    ('Selected Transit Days', 'Least Cost Transit Days'),
                    ('Selected Freight Cost', 'Least Cost Freight Cost'),
                    ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                    ('Selected Total Cost', 'Least Cost Total Cost')
                ]

                for selected_col, least_cost_col in column_pairs:
                    if selected_col in df.columns and least_cost_col in df.columns:
                        df.loc[empty_mask, least_cost_col] = df.loc[empty_mask, selected_col]

                if 'Potential Savings' in df.columns:
                    df.loc[empty_mask, 'Potential Savings'] = 0

                rules_applied.append(f"Empty Data Rule: {count} rows")

        # Rule 3: Negative Savings Rule
        if 'Potential Savings' in df.columns:
            negative_mask = pd.to_numeric(df['Potential Savings'], errors='coerce') < 0
            count = negative_mask.sum()

            if count > 0:
                # Copy selected data to least cost
                column_pairs = [
                    ('Selected Carrier', 'Least Cost Carrier'),
                    ('Selected Service Type', 'Least Cost Service Type'),
                    ('Selected Transit Days', 'Least Cost Transit Days'),
                    ('Selected Freight Cost', 'Least Cost Freight Cost'),
                    ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                    ('Selected Total Cost', 'Least Cost Total Cost')
                ]

                for selected_col, least_cost_col in column_pairs:
                    if selected_col in df.columns and least_cost_col in df.columns:
                        df.loc[negative_mask, least_cost_col] = df.loc[negative_mask, selected_col]

                df.loc[negative_mask, 'Potential Savings'] = 0
                rules_applied.append(f"Negative Savings Rule: {count} rows")

        # Rule 4: TL Carriers Rule
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            tl_mask = (
                df['Selected Carrier'].astype(str).str.upper().isin([c.upper() for c in self.TL_CARRIERS]) |
                df['Least Cost Carrier'].astype(str).str.upper().isin([c.upper() for c in self.TL_CARRIERS])
            )

            count = tl_mask.sum()
            if count > 0:
                # Copy selected to least cost
                column_pairs = [
                    ('Selected Carrier', 'Least Cost Carrier'),
                    ('Selected Service Type', 'Least Cost Service Type'),
                    ('Selected Transit Days', 'Least Cost Transit Days'),
                    ('Selected Freight Cost', 'Least Cost Freight Cost'),
                    ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                    ('Selected Total Cost', 'Least Cost Total Cost')
                ]

                for selected_col, least_cost_col in column_pairs:
                    if selected_col in df.columns and least_cost_col in df.columns:
                        df.loc[tl_mask, least_cost_col] = df.loc[tl_mask, selected_col]

                if 'Potential Savings' in df.columns:
                    df.loc[tl_mask, 'Potential Savings'] = 0

                rules_applied.append(f"TL Carriers Rule: {count} rows")

        # Log all applied rules
        if rules_applied:
            print("✅ Business rules applied:")
            for rule in rules_applied:
                print(f"   • {rule}")
        else:
            print("ℹ️ No business rules needed to be applied")

        return df

    def _calculate_summary_stats(self, df):
        """Calculate summary statistics for the processed data"""
        self.summary_stats = {}

        if 'Potential Savings' in df.columns:
            savings_series = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
            selected_cost_series = pd.to_numeric(df.get('Selected Total Cost', 0), errors='coerce').fillna(0)
            least_cost_series = pd.to_numeric(df.get('Least Cost Total Cost', 0), errors='coerce').fillna(0)

            self.summary_stats = {
                'total_loads': len(df),
                'total_potential_savings': savings_series.sum(),
                'average_savings_per_load': savings_series.mean(),
                'loads_with_savings': (savings_series > 0).sum(),
                'max_single_saving': savings_series.max(),
                'total_selected_cost': selected_cost_series.sum(),
                'total_least_cost': least_cost_series.sum()
            }

            # Calculate percentage savings
            if self.summary_stats['total_selected_cost'] > 0:
                self.summary_stats['percentage_savings'] = (
                    self.summary_stats['total_potential_savings'] / self.summary_stats['total_selected_cost'] * 100
                )
            else:
                self.summary_stats['percentage_savings'] = 0

    def save_to_excel(self, output_file):
        """Save processed data to Excel with professional formatting"""
        if self.processed_data is None:
            raise Exception("No processed data to save")

        # Create workbook
        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = "Basic TMS Report"

        # Add title information
        row = 1
        if self.title_info:
            if 'company_name' in self.title_info:
                cell = ws_data.cell(row=row, column=1, value=f"Company: {self.title_info['company_name']}")
                cell.font = Font(size=12, bold=True)
                row += 1

            if 'date_range' in self.title_info:
                cell = ws_data.cell(row=row, column=1, value=f"Report Period: {self.title_info['date_range']}")
                cell.font = Font(size=11)
                row += 1

            row += 1  # Empty row

        # Add section headers with color coding
        selected_header = ws_data.cell(row=row, column=9, value="Selected Carrier")
        selected_header.font = Font(size=10, bold=True, color="FFFFFF")
        selected_header.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        selected_header.alignment = Alignment(horizontal="center", vertical="center")
        ws_data.merge_cells(f'I{row}:N{row}')

        least_cost_header = ws_data.cell(row=row, column=15, value="Least Cost Carrier")
        least_cost_header.font = Font(size=10, bold=True, color="FFFFFF")
        least_cost_header.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
        least_cost_header.alignment = Alignment(horizontal="center", vertical="center")
        ws_data.merge_cells(f'O{row}:T{row}')

        # Find Potential Savings column
        ps_col = None
        headers = self.processed_data.columns.tolist()
        for i, header in enumerate(headers):
            if header == 'Potential Savings':
                ps_col = i + 1
                break

        if ps_col:
            savings_header = ws_data.cell(row=row, column=ps_col, value="Potential Savings")
            savings_header.font = Font(size=10, bold=True, color="FFFFFF")
            savings_header.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            savings_header.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Add column headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=row, column=col_idx, value=header)
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Add data rows
        for _, data_row in self.processed_data.iterrows():
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_data.cell(row=row, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Apply color coding and currency formatting
                if 9 <= col_idx <= 14:  # Selected Carrier columns
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                elif 15 <= col_idx <= 20:  # Least Cost Carrier columns
                    cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                elif col_idx == ps_col:  # Potential Savings
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    if isinstance(value, (int, float)) and value > 0:
                        cell.font = Font(bold=True, color="28a745")

                # Format currency columns
                header_name = headers[col_idx-1]
                if any(cost_term in header_name for cost_term in ['Cost', 'Savings']):
                    cell.number_format = '"$"#,##0.00'

            row += 1

        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width

        # Add summary sheet
        ws_summary = wb.create_sheet(title="Summary Statistics")
        ws_summary.cell(row=1, column=1, value="Basic TMS Processing Summary").font = Font(size=14, bold=True)

        row = 3
        for key, value in self.summary_stats.items():
            ws_summary.cell(row=row, column=1, value=key.replace('_', ' ').title())
            if isinstance(value, (int, float)):
                if 'cost' in key or 'savings' in key:
                    ws_summary.cell(row=row, column=2, value=f"${value:,.2f}")
                elif 'percentage' in key:
                    ws_summary.cell(row=row, column=2, value=f"{value:.2f}%")
                else:
                    ws_summary.cell(row=row, column=2, value=f"{value:,}")
            else:
                ws_summary.cell(row=row, column=2, value=str(value))
            row += 1

        wb.save(output_file)


class BasicTMSGUI:
    """Simplified GUI for Basic TMS Processing"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Basic TMS Processor")
        self.root.geometry("800x600")
        self.root.configure(bg='#f8f9fa')

        # Initialize processor
        self.processor = BasicTMSProcessor()
        self.input_files = []

        self.setup_gui()

    def setup_gui(self):
        """Setup the main GUI interface"""
        # Main container
        main_frame = tk.Frame(self.root, bg='#ffffff', relief='flat', bd=1)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Title
        title_label = tk.Label(main_frame, text="📊 Basic TMS Processor",
                              font=('Segoe UI', 24, 'bold'), bg='#ffffff', fg='#2d3748')
        title_label.pack(pady=(20, 10))

        subtitle_label = tk.Label(main_frame, text="Simplified TMS report processing with core business logic",
                                 font=('Segoe UI', 12), bg='#ffffff', fg='#4a5568')
        subtitle_label.pack(pady=(0, 30))

        # Input section
        input_frame = tk.LabelFrame(main_frame, text="📁 Input Files", font=('Segoe UI', 12, 'bold'),
                                   bg='#ffffff', fg='#2d3748', relief='flat', bd=1)
        input_frame.pack(fill='x', pady=(0, 20))

        # File selection buttons
        button_frame = tk.Frame(input_frame, bg='#ffffff')
        button_frame.pack(fill='x', padx=20, pady=15)

        self.select_files_btn = tk.Button(button_frame, text="📂 Select Files",
                                         font=('Segoe UI', 11, 'bold'),
                                         bg='#4299e1', fg='white', relief='flat', bd=0,
                                         cursor='hand2', command=self.select_files,
                                         padx=20, pady=10)
        self.select_files_btn.pack(side='left', padx=(0, 10))

        self.clear_files_btn = tk.Button(button_frame, text="🗑️ Clear",
                                        font=('Segoe UI', 11),
                                        bg='#e2e8f0', fg='#4a5568', relief='flat', bd=0,
                                        cursor='hand2', command=self.clear_files,
                                        padx=20, pady=10)
        self.clear_files_btn.pack(side='left')

        # File list
        list_frame = tk.Frame(input_frame, bg='#ffffff')
        list_frame.pack(fill='both', expand=True, padx=20, pady=(0, 15))

        self.file_listbox = tk.Listbox(list_frame, font=('Segoe UI', 10),
                                      bg='#f7fafc', fg='#2d3748', relief='flat', bd=1,
                                      selectbackground='#4299e1', selectforeground='white')
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        self.file_listbox.pack(side='left', fill='both', expand=True)
        self.file_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.file_listbox.yview)

        # Process section
        process_frame = tk.LabelFrame(main_frame, text="⚡ Processing", font=('Segoe UI', 12, 'bold'),
                                     bg='#ffffff', fg='#2d3748', relief='flat', bd=1)
        process_frame.pack(fill='x', pady=(0, 20))

        self.process_btn = tk.Button(process_frame, text="🚀 Process Files",
                                    font=('Segoe UI', 14, 'bold'),
                                    bg='#38a169', fg='white', relief='flat', bd=0,
                                    cursor='hand2', command=self.process_files,
                                    padx=30, pady=15)
        self.process_btn.pack(pady=20)

        # Progress section
        self.progress_frame = tk.LabelFrame(main_frame, text="📈 Progress", font=('Segoe UI', 12, 'bold'),
                                           bg='#ffffff', fg='#2d3748', relief='flat', bd=1)
        self.progress_frame.pack(fill='both', expand=True)

        self.progress_text = tk.Text(self.progress_frame, font=('Consolas', 10),
                                    bg='#1a202c', fg='#e2e8f0', relief='flat', bd=0,
                                    wrap='word', state='disabled')

        progress_scrollbar = tk.Scrollbar(self.progress_frame)
        progress_scrollbar.pack(side='right', fill='y')
        self.progress_text.pack(side='left', fill='both', expand=True, padx=20, pady=15)
        self.progress_text.config(yscrollcommand=progress_scrollbar.set)
        progress_scrollbar.config(command=self.progress_text.yview)

    def log_message(self, message):
        """Add message to progress log"""
        self.progress_text.config(state='normal')
        self.progress_text.insert('end', f"{datetime.now().strftime('%H:%M:%S')} {message}\n")
        self.progress_text.see('end')
        self.progress_text.config(state='disabled')
        self.root.update()

    def select_files(self):
        """Select input files"""
        files = filedialog.askopenfilenames(
            title="Select TMS Excel Files",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if files:
            self.input_files = list(files)
            self.update_file_list()
            self.log_message(f"📁 Selected {len(files)} file(s)")

    def clear_files(self):
        """Clear selected files"""
        self.input_files = []
        self.update_file_list()
        self.log_message("🗑️ Cleared file selection")

    def update_file_list(self):
        """Update the file list display"""
        self.file_listbox.delete(0, tk.END)
        for file_path in self.input_files:
            self.file_listbox.insert(tk.END, os.path.basename(file_path))

    def process_files(self):
        """Process the selected files"""
        if not self.input_files:
            messagebox.showwarning("No Files", "Please select files to process first.")
            return

        # Disable process button during processing
        self.process_btn.config(state='disabled', text="Processing...", bg='#a0aec0')

        # Run processing in thread to prevent GUI freezing
        thread = threading.Thread(target=self._process_files_thread)
        thread.daemon = True
        thread.start()

    def _process_files_thread(self):
        """Process files in separate thread"""
        try:
            processed_files = []
            total_loads = 0
            total_savings = 0.0

            for i, input_file in enumerate(self.input_files, 1):
                self.log_message(f"🔄 Processing file {i}/{len(self.input_files)}: {os.path.basename(input_file)}")

                # Process the file
                self.processor.process_file(input_file)

                # Generate output filename
                input_dir = os.path.dirname(input_file)
                input_name = os.path.splitext(os.path.basename(input_file))[0]
                output_file = os.path.join(input_dir, f"{input_name}_BASIC_PROCESSED.xlsx")

                # Save the file
                self.processor.save_to_excel(output_file)
                processed_files.append(output_file)

                # Update totals
                stats = self.processor.summary_stats
                total_loads += stats.get('total_loads', 0)
                total_savings += stats.get('total_potential_savings', 0)

                self.log_message(f"✅ Saved: {os.path.basename(output_file)}")
                self.log_message(f"📊 Loads: {stats.get('total_loads', 0):,}, "
                               f"Savings: ${stats.get('total_potential_savings', 0):,.2f}")

            # Show completion summary
            self.root.after(0, lambda: self._show_completion_summary(processed_files, total_loads, total_savings))

        except Exception as e:
            self.log_message(f"❌ Error: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Processing Error", f"Error processing files:\n{str(e)}"))
        finally:
            # Re-enable process button
            self.root.after(0, self._enable_process_button)

    def _enable_process_button(self):
        """Re-enable the process button"""
        self.process_btn.config(state='normal', text="🚀 Process Files", bg='#38a169')

    def _show_completion_summary(self, processed_files, total_loads, total_savings):
        """Show completion summary"""
        files_list = "\n".join([os.path.basename(f) for f in processed_files])

        messagebox.showinfo("Processing Complete",
            f"🎉 Successfully processed {len(processed_files)} file(s)!\n\n"
            f"📊 Total Loads: {total_loads:,}\n"
            f"💵 Total Potential Savings: ${total_savings:,.2f}\n\n"
            f"📁 Processed files:\n{files_list}")

    def run(self):
        """Start the GUI application"""
        self.root.mainloop()


def main():
    """Main entry point"""
    app = BasicTMSGUI()
    app.run()


if __name__ == "__main__":
    main()
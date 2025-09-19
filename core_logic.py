#!/usr/bin/env python3
"""
Core TMS Business Logic - Modular Version for Easy Integration

This module contains the core business rules for TMS report processing.
Can be easily imported into other projects or adapted by AI assistants.

Usage Examples:
    # Simple usage
    processor = TMSProcessor()
    df = processor.process_excel_file("report.xlsx")
    processor.save_to_excel(df, "output.xlsx")

    # Custom usage
    processor = TMSProcessor()
    df = processor.load_data("report.xlsx")
    df = processor.apply_business_rules(df)
    processor.save_to_excel(df, "output.xlsx")
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class TMSProcessor:
    """
    Core TMS Processing Logic - Easy to adapt and extend

    Key Features:
    - Modular business rules that can be applied individually
    - Clean data loading and saving methods
    - Professional Excel formatting
    - Comprehensive error handling
    """

    def __init__(self):
        """Initialize processor with default settings"""
        # File structure settings (easily customizable)
        self.HEADER_ROW = 7          # Row 8 in Excel (0-indexed)
        self.DATA_START_ROW = 10     # Row 11 in Excel (0-indexed)

        # TL carriers requiring special processing
        self.TL_CARRIERS = {
            'LANDSTAR RANGER INC',
            'SMARTWAY TRANSPORTATION INC',
            'SMARTWAY CORPORATION INC'
        }

        # Results storage
        self.processed_data = None
        self.summary_stats = {}
        self.title_info = {}

    def process_excel_file(self, file_path):
        """
        Complete processing pipeline - one method does everything

        Args:
            file_path (str): Path to Excel file

        Returns:
            pandas.DataFrame: Processed data ready for output
        """
        print(f"🔄 Processing: {os.path.basename(file_path)}")

        # Step 1: Load data
        df = self.load_data(file_path)

        # Step 2: Apply business rules
        df = self.apply_business_rules(df)

        # Step 3: Calculate summary stats
        self.calculate_summary_stats(df)

        # Step 4: Sort by destination
        if 'Destination City' in df.columns:
            df = df.sort_values('Destination City', na_position='last')

        self.processed_data = df
        return df

    def load_data(self, file_path):
        """
        Load and clean Excel data

        Args:
            file_path (str): Path to Excel file

        Returns:
            pandas.DataFrame: Clean data with proper column names
        """
        # Read Excel file
        df_raw = pd.read_excel(file_path, sheet_name=0, header=None)

        # Extract title information
        self.title_info = self._extract_title_info(df_raw)

        # Get headers and data
        headers = df_raw.iloc[self.HEADER_ROW].fillna('').astype(str).tolist()
        data_df = df_raw.iloc[self.DATA_START_ROW:].copy()
        data_df.columns = headers

        # Handle PS -> Potential Savings column mapping
        if 'PS' in data_df.columns:
            data_df = data_df.rename(columns={'PS': 'Potential Savings'})

        # Clean the data
        data_df = self._clean_data(data_df)

        # Calculate Potential Savings if missing
        data_df = self._calculate_potential_savings(data_df)

        return data_df

    def apply_business_rules(self, df):
        """
        Apply all TMS business rules

        Args:
            df (pandas.DataFrame): Input data

        Returns:
            pandas.DataFrame: Data with business rules applied
        """
        df = df.copy()

        print("📋 Applying business rules...")

        # Rule 1: Same Carrier Rule
        df = self._apply_same_carrier_rule(df)

        # Rule 2: Empty Least Cost Data Rule
        df = self._apply_empty_data_rule(df)

        # Rule 3: Negative Savings Rule
        df = self._apply_negative_savings_rule(df)

        # Rule 4: TL Carriers Rule
        df = self._apply_tl_carriers_rule(df)

        return df

    def _apply_same_carrier_rule(self, df):
        """Rule 1: Set PS to 0 when selected carrier = least cost carrier"""
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            mask = (
                (df['Selected Carrier'].astype(str) == df['Least Cost Carrier'].astype(str)) &
                (df['Selected Carrier'].notna()) &
                (df['Least Cost Carrier'].notna()) &
                (df['Selected Carrier'].astype(str) != '') &
                (df['Least Cost Carrier'].astype(str) != '')
            )

            count = mask.sum()
            if count > 0 and 'Potential Savings' in df.columns:
                df.loc[mask, 'Potential Savings'] = 0
                print(f"✅ Same Carrier Rule: {count} rows")

        return df

    def _apply_empty_data_rule(self, df):
        """Rule 2: Copy selected data when least cost data is missing"""
        if 'Least Cost Carrier' in df.columns:
            mask = (
                df['Least Cost Carrier'].isna() |
                (df['Least Cost Carrier'].astype(str) == '') |
                (df['Least Cost Carrier'].astype(str) == 'nan')
            )

            count = mask.sum()
            if count > 0:
                self._copy_selected_to_least_cost(df, mask)
                if 'Potential Savings' in df.columns:
                    df.loc[mask, 'Potential Savings'] = 0
                print(f"✅ Empty Data Rule: {count} rows")

        return df

    def _apply_negative_savings_rule(self, df):
        """Rule 3: Fix negative potential savings"""
        if 'Potential Savings' in df.columns:
            ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
            mask = ps_numeric < 0
            count = mask.sum()

            if count > 0:
                self._copy_selected_to_least_cost(df, mask)
                df.loc[mask, 'Potential Savings'] = 0
                print(f"✅ Negative Savings Rule: {count} rows")

        return df

    def _apply_tl_carriers_rule(self, df):
        """Rule 4: Special handling for TL carriers"""
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            mask = (
                df['Selected Carrier'].astype(str).str.upper().isin([c.upper() for c in self.TL_CARRIERS]) |
                df['Least Cost Carrier'].astype(str).str.upper().isin([c.upper() for c in self.TL_CARRIERS])
            )

            count = mask.sum()
            if count > 0:
                self._copy_selected_to_least_cost(df, mask)
                if 'Potential Savings' in df.columns:
                    df.loc[mask, 'Potential Savings'] = 0
                print(f"✅ TL Carriers Rule: {count} rows")

        return df

    def _copy_selected_to_least_cost(self, df, mask):
        """Helper: Copy selected carrier data to least cost columns"""
        column_pairs = [
            ('Selected Carrier', 'Least Cost Carrier'),
            ('Selected Service Type', 'Least Cost Service Type'),
            ('Selected Transit Days', 'Least Cost Transit Days'),
            ('Selected Freight Cost', 'Least Cost Freight Cost'),
            ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
            ('Selected Total Cost', 'Least Cost Total Cost')
        ]

        for selected_col, least_cost_col in column_pairs:
            if selected_col in df.columns and least_cost_col in df.columns:
                df.loc[mask, least_cost_col] = df.loc[mask, selected_col]

    def _clean_data(self, df):
        """Clean and validate data types"""
        df = df.copy()

        # Remove rows with missing Load No.
        if 'Load No.' in df.columns:
            df = df.dropna(subset=['Load No.'])
            df = df[df['Load No.'].astype(str).str.strip() != '']
            df = df[df['Load No.'].astype(str) != 'nan']

        # Clean numeric columns
        numeric_columns = [
            'Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
            'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost',
            'Potential Savings'
        ]

        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Clean string columns
        string_columns = [
            'Selected Carrier', 'Selected Service Type', 'Least Cost Carrier', 'Least Cost Service Type'
        ]

        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')

        # Remove mostly empty rows
        df = df.dropna(thresh=5).reset_index(drop=True)

        return df

    def _calculate_potential_savings(self, df):
        """Calculate Potential Savings if missing"""
        df = df.copy()

        # If PS already exists and has valid data, keep it
        if 'Potential Savings' in df.columns:
            ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce')
            if ps_numeric.notna().sum() > 0:
                return df

        # Calculate PS from cost difference
        if 'Selected Total Cost' in df.columns and 'Least Cost Total Cost' in df.columns:
            selected_cost = pd.to_numeric(df['Selected Total Cost'], errors='coerce').fillna(0)
            least_cost = pd.to_numeric(df['Least Cost Total Cost'], errors='coerce').fillna(0)
            df['Potential Savings'] = selected_cost - least_cost

            # Handle cases where least cost is 0
            mask_zero_least = least_cost == 0
            df.loc[mask_zero_least, 'Potential Savings'] = 0
        else:
            df['Potential Savings'] = 0

        return df

    def _extract_title_info(self, df_raw):
        """Extract title and company information from Excel header"""
        title_info = {}

        try:
            # Company name (row 4, column B)
            if len(df_raw) > 3 and len(df_raw.columns) > 1:
                company = df_raw.iloc[3, 1]
                if pd.notna(company):
                    title_info['company_name'] = str(company)

            # Date range (row 6, column B)
            if len(df_raw) > 5 and len(df_raw.columns) > 1:
                date_range = df_raw.iloc[5, 1]
                if pd.notna(date_range):
                    title_info['date_range'] = str(date_range)
        except Exception:
            pass

        return title_info

    def calculate_summary_stats(self, df):
        """Calculate comprehensive summary statistics"""
        if df.empty:
            self.summary_stats = {
                'total_loads': 0,
                'total_potential_savings': 0,
                'average_savings_per_load': 0,
                'loads_with_savings': 0,
                'total_selected_cost': 0,
                'total_least_cost': 0,
                'percentage_savings': 0
            }
            return

        # Calculate all stats
        total_loads = len(df)
        total_selected_cost = pd.to_numeric(df.get('Selected Total Cost', 0), errors='coerce').fillna(0).sum()
        total_least_cost = pd.to_numeric(df.get('Least Cost Total Cost', 0), errors='coerce').fillna(0).sum()
        total_potential_savings = pd.to_numeric(df.get('Potential Savings', 0), errors='coerce').fillna(0).sum()

        ps_numeric = pd.to_numeric(df.get('Potential Savings', 0), errors='coerce').fillna(0)
        loads_with_savings = (ps_numeric > 0).sum()

        percentage_savings = (total_potential_savings / total_selected_cost * 100) if total_selected_cost > 0 else 0
        average_savings_per_load = total_potential_savings / total_loads if total_loads > 0 else 0

        self.summary_stats = {
            'total_loads': total_loads,
            'total_potential_savings': total_potential_savings,
            'average_savings_per_load': average_savings_per_load,
            'loads_with_savings': loads_with_savings,
            'total_selected_cost': total_selected_cost,
            'total_least_cost': total_least_cost,
            'percentage_savings': percentage_savings
        }

    def save_to_excel(self, df, output_file):
        """
        Save processed data to Excel with professional formatting

        Args:
            df (pandas.DataFrame): Processed data
            output_file (str): Output file path
        """
        if df is None or df.empty:
            raise Exception("No data to save")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Basic TMS Report"

        # Add title information
        row = 1
        if self.title_info:
            if 'company_name' in self.title_info:
                cell = ws.cell(row=row, column=1, value=f"Company: {self.title_info['company_name']}")
                cell.font = Font(size=12, bold=True)
                row += 1

            if 'date_range' in self.title_info:
                cell = ws.cell(row=row, column=1, value=f"Report Period: {self.title_info['date_range']}")
                cell.font = Font(size=11)
                row += 1

            row += 1  # Empty row

        # Add headers with color coding
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Add data rows
        for _, data_row in df.iterrows():
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Format currency columns
                header_name = headers[col_idx-1]
                if any(cost_term in header_name for cost_term in ['Cost', 'Savings']):
                    cell.number_format = '"$"#,##0.00'

            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file)
        print(f"💾 Saved: {output_file}")


# Example usage for testing
if __name__ == "__main__":
    # Simple example
    processor = TMSProcessor()

    # Example file path (replace with actual file)
    # df = processor.process_excel_file("sample_report.xlsx")
    # processor.save_to_excel(df, "output_processed.xlsx")

    print("✅ Core TMS Logic module ready for use!")
    print("📖 Import this module into your project:")
    print("   from core_logic import TMSProcessor")